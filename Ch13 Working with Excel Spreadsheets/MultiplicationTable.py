import openpyxl
import sys

if len(sys.argv) == 2 and int(sys.argv[1]) >= 0:
    inputNum = int(sys.argv[1])
    wb = openpyxl.Workbook()
    sheet = wb.active
    for row in range(1,inputNum+1):
        for col in range(1,inputNum+1):
            # ignore the base 1,1 cell
            if row == 1 and col == 1:
                continue 
            # writing the "header numbers" of the table in the first row and column
            if row == 1:
                sheet.cell(row=row, column=col).value = col
            elif col == 1:
                sheet.cell(row=row, column=col).value = row
            # performing and writing the result of the actual math in all other cases
            else:
                sheet.cell(row=row, column=col).value = row*col
    
    wb.save("Multiplication Table up to "+str(inputNum)+".xlsx")

else: print("Invalid input given! Please pass a non-negative integer.")
